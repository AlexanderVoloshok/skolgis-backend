import json
import pandas as pd
import geopandas as gpd
from typing import Any, Dict, List, Tuple, Optional, Union
from sqlalchemy import and_, or_, not_, true, false, Column, String, DateTime, Numeric, text, MetaData, Table
from sqlalchemy.sql.expression import BinaryExpression
from openpyxl import load_workbook
from openpyxl.styles import Font
from src.config import ENGINE

def parse_filters(raw) -> List[Dict[str, Any]]:
    if raw is None:
        return []
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            raise ValueError("filters must be valid JSON")
    if isinstance(raw, list):
        return raw
    raise ValueError("filters must be an array")

def build_where(filters_raw: Any, columns: Dict[str, Column[Any]]) -> BinaryExpression:

    rules = parse_filters(filters_raw)
    clauses: List[BinaryExpression] = []

    for r in rules:
        if not isinstance(r, dict):
            continue
        col = r.get("field")
        op = r.get("op")
        value = r.get("value")
        inclusive = r.get("inclusive", True)

        if col not in columns:
            raise ValueError(f"Forbidden field: {col}")

        col = columns[col]

        # BETWEEN / NOT BETWEEN
        if op in ("BETWEEN", "NOT BETWEEN"):
            if not isinstance(value, list):
                raise ValueError("BETWEEN requires [min, max]")
            a, b = (value + [None, None])[:2]
            # нормализуем порядок, если оба заданы и сравнимы
            try:
                if a is not None and b is not None and a > b:
                    a, b = b, a
            except Exception:
                pass

            if a is None and b is None:
                continue

            if a is not None and b is not None:
                expr = col.between(a, b) if inclusive else and_(col > a, col < b)
                if op == "NOT BETWEEN":
                    expr = not_(expr) if inclusive else or_(col <= a, col >= b)
                clauses.append(expr)
                continue
            if a is not None:
                expr = (col >= a) if inclusive else (col > a)
                if op == "NOT BETWEEN":
                    expr = (col < a) if inclusive else (col <= a)
                clauses.append(expr)
                continue
            if b is not None:
                expr = (col <= b) if inclusive else (col < b)
                if op == "NOT BETWEEN":
                    expr = (col > b) if inclusive else (col >= b)
                clauses.append(expr)
                continue

        # IN
        elif op == "IN":
            if not isinstance(value, list):
                raise ValueError("IN requires array value")
            clauses.append(false() if len(value) == 0 else col.in_(value))

        # ILIKE
        elif op == "ILIKE":
            if value is None:  # игнорируем пустую маску
                continue
            clauses.append(col.ilike(str(value)))

        # базовые сравнения
        elif op in ("=", "!=", ">", ">=", "<", "<="):
            if op == "=":   clauses.append(col == value)
            if op == "!=":  clauses.append(col != value)
            if op == ">":   clauses.append(col >  value)
            if op == ">=":  clauses.append(col >= value)
            if op == "<":   clauses.append(col <  value)
            if op == "<=":  clauses.append(col <= value)

    #clauses.append(columns['geom'] != None)

    return and_(*clauses) if clauses else true()

def parse_order(raw: Optional[str]) -> List[Tuple[str, str]]:
    if not raw:
        return []
    out: List[Tuple[str, str]] = []
    for part in str(raw).split(","):
        part = part.strip()
        if not part: 
            continue
        col, dir_ = (part.split(":", 1) + ["asc"])[:2]
        if col == 'None':
            continue
        col = col.strip()
        dir_ = dir_.strip().lower()
        if dir_ not in ("asc", "desc"):
            raise ValueError("Bad order direction (asc/desc)")
        out.append((col, dir_))
    return out

def resolve_type(kind: str, length: int = None):
    """Маппинг 'строка/дата/число' -> тип SQLAlchemy."""
    kind = kind.lower()

    # строка
    if kind in ("текст", "string", "str", "text"):
        return String(length or 150)

    # дата
    if kind in ("дата", "date"):
        return DateTime()

    # число (по умолчанию целое)
    if kind in ("число", "number", "int", "integer"):
        return Numeric()

    raise ValueError(f"Неизвестный тип колонки: {kind}")

def add_category_totals(
    gdf: Union[gpd.GeoDataFrame, pd.DataFrame],
    category_col: str,
    total_label_col: Optional[str] = None,
    grand_total_text: str = "ИТОГО",
    exclude_sum_cols: Optional[List[str]] = ['project_id', 'id'],
    keep_geometry: bool = False,
) -> pd.DataFrame:
    """
    Добавляет строки 'ИТОГО' после каждой категории и общий итог в конце.

    Args:
        gdf: исходный GeoDataFrame/DataFrame
        category_col: колонка с категорией объектов
        total_label_col: в какую колонку писать 'ИТОГО'
                         если None, текст пишется в category_col
        total_text: подпись для итогов по категории
        grand_total_text: подпись для общего итога
        exclude_sum_cols: числовые колонки, которые НЕ нужно суммировать
        keep_geometry: оставлять ли geometry в результате

    Returns:
        pd.DataFrame с добавленными строками итогов
    """
    df = gdf.copy()

    if not keep_geometry and "geometry" in df.columns:
        df = df.drop(columns=["geometry"])

    exclude_sum_cols = set(exclude_sum_cols or [])

    # Какие колонки суммировать
    numeric_cols = [
        col for col in df.select_dtypes(include="number").columns
        if col not in exclude_sum_cols
    ]

    # Куда писать слово "ИТОГО"
    label_col = total_label_col or category_col

    # Сохраняем исходный порядок категорий
    categories = df[category_col].drop_duplicates().tolist()
    if None in categories:
        categories.remove(None)

    parts = []

    for category in categories:
        group = df[df[category_col] == category].copy()
        parts.append(group)

        total_row = {col: None for col in df.columns}

        # Подпись строки
        total_row[label_col] = f"ИТОГО {category}"

        # Чтобы было видно, к какой категории относится итог
        if label_col != category_col:
            total_row[category_col] = category

        # Суммы только по числовым колонкам
        for col in numeric_cols:
            total_row[col] = group[col].sum()

        parts.append(pd.DataFrame([total_row], columns=df.columns))

    # Общий итог
    grand_total_row = {col: None for col in df.columns}
    grand_total_row[label_col] = grand_total_text

    for col in numeric_cols:
        grand_total_row[col] = df[col].sum()

    parts.append(pd.DataFrame([grand_total_row], columns=df.columns))

    result = pd.concat(parts, ignore_index=True)
    return result

def bold_total_rows(xlsx_path: str, sheet_name: str, label_col_name: str) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]]

    label_col_idx = headers.index(label_col_name) + 1

    bold_font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=label_col_idx).value
        if isinstance(value, str) and value.startswith("ИТОГО"):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = bold_font

    wb.save(xlsx_path)

def get_refresh_projects_view_query(columns: List[Dict]):
    """
    Генерирует CREATE OR REPLACE VIEW для projects_full.

    Логика:
    - все колонки берутся из projects_attrs динамически
    - в 1-м SELECT идут реальные p.<col>
    - во 2-м SELECT идут либо override-выражения, либо NULL::<type> AS <col>
    """

    attrs_table: str = "projects_attrs"
    buildings_table: str = "main_buildings"

    # Если есть колонки, которые во 2-м SELECT надо брать не как NULL,
    # а из main_buildings — задаём их здесь.
    buildings_select_overrides = {
        "geom": "mb.geom as geom",
        "fids": "mb.id::text AS fids",
        "floors": "mb.floors::text AS floors",
    }

    # 1) Первый SELECT:
    #    id = p.project_id AS id
    #    затем все колонки p в естественном порядке
    #    затем geom и fids из агрегата
    project_attrs_select_parts = [
        "p.id AS id",
        "b.geom",
        "b.fids",
        "p.id AS project_id",
    ]

    project_attrs_select_parts.extend([f"p.{col}" for col in columns.keys() if col != 'id'])
    project_attrs_select_parts.append("(dfp.code || ' '::text) || ds.code AS func_purpose_stage")

    first_select_sql = " SELECT " + ",\n        ".join(project_attrs_select_parts) + f"""
        FROM skolkovo_layers.{attrs_table} p
        LEFT JOIN b_by_project b ON b.project_id = p.id
        LEFT JOIN skolkovo_general.dict_func_purpose dfp on p.func_purpose = dfp.name
        LEFT JOIN skolkovo_general.dict_stage ds on p.stage = ds.name
        WHERE b.geom is not null OR p.name is not null OR p.func_purpose is not null;
    """

    sql = f"""
        CREATE VIEW skolkovo_layers.projects_full
        AS
        WITH b_by_project AS (
            SELECT
                mb.project_id,
                ST_CollectionExtract(st_collect(mb.geom), 3) AS geom,
                string_agg(mb.id::text, ','::text) AS fids
            FROM skolkovo_layers.{buildings_table} mb
            WHERE mb.project_id IS NOT NULL
            GROUP BY mb.project_id
        )
        {first_select_sql}
        
    """.strip()

    return text(sql)


def get_refresh_parcels_view_query():
    sql = """CREATE OR REPLACE VIEW skolkovo_layers.parcels
        AS SELECT p.*, (dfp.code::text || ' '::text) || ds.code::text AS func_purpose_stage
        FROM skolkovo_layers.layer_32 p
             LEFT JOIN skolkovo_general.dict_func_purpose dfp ON p.objects_type = dfp.name::text
             LEFT JOIN skolkovo_general.dict_stage ds ON p.status = ds.name::text;
    """.strip()

    return text(sql)


def reflect_layer_table(name: str):
    meta = MetaData()
    return Table(
        name,
        meta,
        schema="skolkovo_layers",
        autoload_with=ENGINE,
        extend_existing=True,
    )
